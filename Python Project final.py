import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Loading the excel sheet
book = openpyxl.load_workbook('C:/Users/LVST/OneDrive/Desktop/Python Project/project.xlsx')

# Choose the sheet
sheet = book['Sheet1']

# Counting number of rows / students
r = sheet.max_row

# Variable for looping for input
resp = 1

# Counting number of columns / subjects
c = sheet.max_column

# Extracting the name of the subjects
sub = []
for i in range(3, c + 1):
    s1 = sheet.cell(row=1, column=i).value
    sub.append(s1)

# Choose another sheet
sheet1 = book['Sheet2']
r1 = sheet1.max_row

# Warning messages
m1 = "Warning!!! From today onwards you cannot take leave for "
m2 = "You have lack of attendance in "
m3 = "The following students have lack of attendance in your subject: "

def savefile():
    book.save(r'C:/Users/LVST/OneDrive/Desktop/Python Project/project1.xlsx')

def check(no_of_days, row_num, b):
    global l1
    global l2
    global l3

    for student in range(len(row_num)):
        if no_of_days[student] == 5:
            l1.append(sheet.cell(row=row_num[student], column=2).value)
        elif no_of_days[student] > 5:
            l2 += str(sheet.cell(row=row_num[student], column=1).value) + ','
            l3.append(sheet.cell(row=row_num[student], column=2).value)

    count = 0
    if l1:
        messag1 = m1 + sub[b - 3] + ' class.'
        mailstu(l1, messag1)
        count = 1

    if l2 and l3:
        messag2 = m2 + sub[b - 3] + "!!!"
        messag3 = m3 + l2
        mailstu(l3, messag2)  # Mail to students
        count = 1
        subjectname = sheet.cell(row=1, column=b).value
        for i in range(2, r1 + 1):
            if subjectname == sheet1.cell(row=i, column=1).value:
                staff = sheet1.cell(row=i, column=2).value
        mailstaff(staff, messag3)

    if count == 1:
        print("Mail sent to students")

def send_email(from_id, pwd, to_id, subject, body):
    with smtplib.SMTP('smtp.gmail.com', 587) as s:
        s.starttls()
        s.login(from_id, pwd)
        message = MIMEMultipart()
        message['From'] = from_id
        message['To'] = to_id
        message['Subject'] = subject
        message.attach(MIMEText(body, 'plain'))
        s.sendmail(from_id, to_id, message.as_string())
    

def mailstu(email_list, msg):
    from_id = 'varshith270304@gmail.com'
    pwd = 'uyyn ibjv tvss kwfo'
    subject = 'Attendance report'
    for email in email_list:
        send_email(from_id, pwd, email, subject, msg)

def mailstaff(mail_id, msg):
    from_id = 'varshith270304@gmail.com'
    pwd = 'uyyn ibjv tvss kwfo'
    subject = 'Lack of attendance report'
    send_email(from_id, pwd, mail_id, subject, msg)
    print("Mail sent to staff")

while resp == 1:
    for i in range(len(sub)):
        print(i + 1, '.', sub[i], '\n')

    # List of students to remind
    l1 = []

    # To concatenate list of roll numbers with lack of attendance
    l2 = ""

    # List of roll numbers with lack of attendance
    l3 = []


    y = int(input("Enter subject number: "))
    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees >= 1:
        x = list(map(int, input('Enter Roll No\'s of absentees: ').split()))
    else:
        print("Surprised to hear that!!")
        resp = int(input('1. Want to Enter the absentees details of another subject\n2. Exit\nEnter your choice: '))
        continue

    row_num = []
    no_of_days = []

    for student in x:
        for i in range(2, r + 1):
            j = y + 2
            if sheet.cell(row=i, column=1).value == student:
                m = sheet.cell(row=i, column=j).value
                m += 1
                sheet.cell(row=i, column=3).value = m
                savefile()
                no_of_days.append(m)
                row_num.append(i)

    print("Absentees Data is Saved Successfully!!")
    check(no_of_days, row_num, y + 2)
    resp = int(input('1. Want to Enter the absentees details of another subject\n2. Exit\nEnter your choice: '))
